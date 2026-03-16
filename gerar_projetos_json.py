from pathlib import Path
import json
import re
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
PROJETOS_DIR = BASE_DIR / "assets" / "projetos"
JSON_PATH = BASE_DIR / "projetos.json"
XLSX_PATH = BASE_DIR / "projetos_metadata.xlsx"

TIPOS_VALIDOS = {"livro", "hq", "revista"}


def limpar(v):
    if v is None:
        return ""
    return str(v).strip().lower()


def slug_para_titulo(slug):
    texto = slug.replace("-", " ").replace("_", " ")
    texto = re.sub(r"(\d+)$", r" \1", texto)
    return " ".join(p.capitalize() for p in texto.split())


def carregar_excel():

    wb = load_workbook(XLSX_PATH)
    ws = wb["metadata"]

    headers = [limpar(c.value) for c in ws[1]]
    idx = {nome: i for i, nome in enumerate(headers)}

    metadata = {}

    for row in ws.iter_rows(min_row=2, values_only=True):

        slug = limpar(row[idx["slug"]]) if "slug" in idx else ""
        cliente = limpar(row[idx["cliente"]]) if "cliente" in idx else ""
        tipo = limpar(row[idx["tipo"]]) if "tipo" in idx else ""

        if not slug or not cliente or not tipo:
            continue

        chave = (cliente, tipo, slug)

        metadata[chave] = {
            "titulo": row[idx["titulo"]] if "titulo" in idx else "",
            "subtitulo": row[idx["subtitulo"]] if "subtitulo" in idx else "",
            "descricao": row[idx["descricao"]] if "descricao" in idx else "",
            "tags": row[idx["tags"]] if "tags" in idx else "",
            "ano": row[idx["ano"]] if "ano" in idx else "",
            "ordem": row[idx["ordem"]] if "ordem" in idx else "",
            "ativo": limpar(row[idx["ativo"]]) if "ativo" in idx else "sim",
        }

    return wb, ws, metadata


def coletar_imagens(pasta, slug):

    thumb = pasta / f"{slug}_thumb.jpg"

    imagens = sorted(
        p for p in pasta.glob(f"{slug}_*.jpg")
        if p.name != f"{slug}_thumb.jpg"
    )

    thumb_rel = str(thumb.relative_to(BASE_DIR)).replace("\\", "/") if thumb.exists() else ""
    imagens_rel = [str(p.relative_to(BASE_DIR)).replace("\\", "/") for p in imagens]

    return thumb_rel, imagens_rel


def adicionar_linha_excel(ws, cliente, tipo, slug):

    ws.append([
        slug,
        cliente,
        tipo,
        slug_para_titulo(slug),
        "",
        "",
        "",
        "",
        "",
        "sim"
    ])


def gerar_catalogo():

    wb, ws, metadata = carregar_excel()

    projetos = []

    novos = 0

    for pasta_cliente in sorted(PROJETOS_DIR.iterdir()):

        if not pasta_cliente.is_dir():
            continue

        cliente = limpar(pasta_cliente.name)

        for pasta_tipo in sorted(pasta_cliente.iterdir()):

            if not pasta_tipo.is_dir():
                continue

            tipo = limpar(pasta_tipo.name)

            if tipo not in TIPOS_VALIDOS:
                continue

            for pasta_projeto in sorted(pasta_tipo.iterdir()):

                if not pasta_projeto.is_dir():
                    continue

                slug = limpar(pasta_projeto.name)

                chave = (cliente, tipo, slug)

                if chave not in metadata:
                    adicionar_linha_excel(ws, cliente, tipo, slug)
                    novos += 1
                    metadata[chave] = {}

                thumb, imagens = coletar_imagens(pasta_projeto, slug)

                if not thumb and not imagens:
                    continue

                meta = metadata.get(chave, {})

                if limpar(meta.get("ativo", "sim")) == "nao":
                    continue

                tags = []
                if meta.get("tags"):
                    tags = [t.strip().lower() for t in str(meta["tags"]).split(",") if t.strip()]

                projeto = {
                    "slug": slug,
                    "titulo": meta.get("titulo") or slug_para_titulo(slug),
                    "subtitulo": meta.get("subtitulo", ""),
                    "descricao": meta.get("descricao", ""),
                    "tipo": tipo,
                    "cliente": cliente,
                    "tags": tags,
                    "ano": meta.get("ano", ""),
                    "ordem": meta.get("ordem", ""),
                    "thumb": thumb,
                    "imagens": imagens
                }

                projetos.append(projeto)

    if novos > 0:
        wb.save(XLSX_PATH)

        if novos > 0:
            wb.save(XLSX_PATH)

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(projetos, f, ensure_ascii=False, indent=2)

    print("Projetos JSON:", len(projetos))
    print("Linhas novas criadas no Excel:", novos)


if __name__ == "__main__":
    gerar_catalogo()